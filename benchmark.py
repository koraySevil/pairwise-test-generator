#!/usr/bin/env python3
"""
benchmark.py – Benchmark suite for pairwise.py

Runs all standard benchmark configurations from the pairwise testing
literature and prints a comparison table.

Usage:
  python benchmark.py                    # greedy only (fast)
  python benchmark.py --exact            # enable exact solver where T <= T_MAX
  python benchmark.py --exact --T_MAX 100000 --exact_time_limit 30

Each benchmark is defined by a compact notation like "3^4" meaning
4 parameters with 3 values each, or "4^15 3^17 2^29" meaning
15 params with 4 values, 17 with 3, 29 with 2.
"""

from __future__ import annotations

import argparse
import os
import sys
import tempfile
import time
from typing import List, Tuple

import openpyxl

# Ensure pairwise module is importable from same directory
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import pairwise


# ---------------------------------------------------------------------------
# Benchmark definitions
# ---------------------------------------------------------------------------

# Each entry: (label, [(value_count, param_count), ...])
#   e.g., "3^4" → [(3, 4)]  means 4 parameters each with 3 values
#   "4^15 3^17 2^29" → [(4,15), (3,17), (2,29)]

BENCHMARKS: List[Tuple[str, List[Tuple[int, int]]]] = [
    ("3^4",                [(3, 4)]),
    ("3^13",               [(3, 13)]),
    ("4^15 3^17 2^29",     [(4, 15), (3, 17), (2, 29)]),
    ("4^1 3^39 2^35",      [(4, 1), (3, 39), (2, 35)]),
    ("2^100",              [(2, 100)]),
    ("10^20",              [(10, 20)]),
    ("4^10",               [(4, 10)]),
    ("4^20",               [(4, 20)]),
    ("4^30",               [(4, 30)]),
    ("4^40",               [(4, 40)]),
    ("4^50",               [(4, 50)]),
    ("4^60",               [(4, 60)]),
    ("4^70",               [(4, 70)]),
    ("4^80",               [(4, 80)]),
    ("4^90",               [(4, 90)]),
    ("4^100",              [(4, 100)]),
]

# Reference results from the literature (IPO, TConfig, AETG, CTS)
# Acquired from https://www.sciencedirect.com/science/article/pii/S0012365X0400130X#SEC5
REFERENCE: dict[str, dict[str, int | None]] = {
    "3^4":             {"IPO": 10, "TConfig": 9,   "AETG": 9,   "CTS": 9},
    "3^13":            {"IPO": 20, "TConfig": 15,  "AETG": 15,  "CTS": 15},
    "4^15 3^17 2^29":  {"IPO": 34, "TConfig": 40,  "AETG": 41,  "CTS": 39},
    "4^1 3^39 2^35":   {"IPO": 27, "TConfig": 30,  "AETG": 28,  "CTS": 29},
    "2^100":           {"IPO": 15, "TConfig": 14,  "AETG": None, "CTS": 10},
    "10^20":           {"IPO": 219,"TConfig": 231, "AETG": 180, "CTS": 210},
    "4^10":            {"IPO": 31, "TConfig": 28,  "AETG": None, "CTS": 28},
    "4^20":            {"IPO": 34, "TConfig": 28,  "AETG": None, "CTS": 28},
    "4^30":            {"IPO": 41, "TConfig": 40,  "AETG": None, "CTS": 40},
    "4^40":            {"IPO": 42, "TConfig": 40,  "AETG": None, "CTS": 40},
    "4^50":            {"IPO": 47, "TConfig": 40,  "AETG": None, "CTS": 40},
    "4^60":            {"IPO": 47, "TConfig": 40,  "AETG": None, "CTS": 40},
    "4^70":            {"IPO": 49, "TConfig": 40,  "AETG": None, "CTS": 40},
    "4^80":            {"IPO": 49, "TConfig": 40,  "AETG": None, "CTS": 40},
    "4^90":            {"IPO": 52, "TConfig": 43,  "AETG": None, "CTS": 43},
    "4^100":           {"IPO": 52, "TConfig": 43,  "AETG": None, "CTS": 43},
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def expand_params(spec: List[Tuple[int, int]]) -> List[List[str]]:
    """
    Expand a benchmark spec into a values list.
    E.g., [(3, 4)] → 4 parameters each with values ["v1", "v2", "v3"].
    """
    values: List[List[str]] = []
    param_idx = 0
    for val_count, param_count in spec:
        for _ in range(param_count):
            values.append([f"v{i}" for i in range(1, val_count + 1)])
            param_idx += 1
    return values


def create_input_excel(values: List[List[str]], path: str) -> None:
    """Write a benchmark input Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, vals in enumerate(values, 1):
        ws.cell(1, col, value=f"P{col}")
        for row, v in enumerate(vals, 2):
            ws.cell(row, col, value=v)
    wb.save(path)


def fmt_val(v: int | None) -> str:
    """Format a value or '-' for None."""
    return str(v) if v is not None else "-"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Benchmark pairwise.py")
    ap.add_argument("--exact", action="store_true",
                    help="Enable exact solver where T <= T_MAX")
    ap.add_argument("--T_MAX", type=int, default=50_000,
                    help="Cartesian threshold for exact phase (default: 50000)")
    ap.add_argument("--exact_time_limit", type=float, default=None,
                    help="Time limit for exact phase per benchmark (seconds)")
    ap.add_argument("--benchmarks", nargs="*", default=None,
                    help="Run only specific benchmarks by label (e.g. '3^4' '4^10')")
    args = ap.parse_args()

    # Filter benchmarks if requested
    benchmarks = BENCHMARKS
    if args.benchmarks:
        selected = set(args.benchmarks)
        benchmarks = [(l, s) for l, s in BENCHMARKS if l in selected]
        if not benchmarks:
            print(f"No benchmarks matched: {args.benchmarks}")
            sys.exit(1)

    # Table header
    hdr_cols = ["Parameter sizes", "LB", "Greedy(UB)", "Final", "Optimal?",
                "Time(s)", "IPO", "TConfig", "AETG", "CTS"]
    col_widths = [22, 6, 12, 6, 8, 8, 6, 8, 6, 6]

    def fmt_row(vals: list) -> str:
        return " | ".join(str(v).ljust(w) for v, w in zip(vals, col_widths))

    sep = "-+-".join("-" * w for w in col_widths)

    print()
    print(fmt_row(hdr_cols))
    print(sep)

    with tempfile.TemporaryDirectory() as tmpdir:
        for label, spec in benchmarks:
            values = expand_params(spec)
            n = len(values)

            input_path = os.path.join(tmpdir, "input.xlsx")
            output_path = os.path.join(tmpdir, "output.xlsx")
            create_input_excel(values, input_path)

            # Build CLI args
            cli_args = [
                "--input", input_path,
                "--output", output_path,
                "--T_MAX", str(args.T_MAX),
            ]
            if not args.exact:
                cli_args.append("--no-exact")
            if args.exact_time_limit is not None:
                cli_args.extend(["--exact_time_limit", str(args.exact_time_limit)])

            # Run
            t0 = time.perf_counter()

            # Capture stdout
            import io
            from contextlib import redirect_stdout
            buf = io.StringIO()
            try:
                with redirect_stdout(buf):
                    pairwise.main(cli_args)
                elapsed = time.perf_counter() - t0
                output = buf.getvalue()

                # Parse results from stdout
                lb = ub = final = "?"
                optimal = "?"
                for token in output.split():
                    if token.startswith("LB="):
                        lb = token.split("=")[1]
                    elif token.startswith("UB="):
                        ub = token.split("=")[1]
                    elif token.startswith("FINAL_SIZE="):
                        final = token.split("=")[1]
                    elif token.startswith("OPTIMAL_PROVEN="):
                        optimal = token.split("=")[1]

            except Exception as e:
                elapsed = time.perf_counter() - t0
                lb = ub = final = "ERR"
                optimal = str(e)[:20]

            # Reference values
            ref = REFERENCE.get(label, {})

            row = [
                label,
                lb,
                ub,
                final,
                optimal,
                f"{elapsed:.2f}",
                fmt_val(ref.get("IPO")),
                fmt_val(ref.get("TConfig")),
                fmt_val(ref.get("AETG")),
                fmt_val(ref.get("CTS")),
            ]
            print(fmt_row(row))

    print()
    print("Done.")


if __name__ == "__main__":
    main()
