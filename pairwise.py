#!/usr/bin/env python3
"""
pairwise.py – Pairwise (2-way) test-suite generator.

Hybrid strategy:
  1. IPOG greedy algorithm for a fast valid upper bound.
  2. Optional exact CP-SAT solver (OR-Tools) for the true minimal suite.

Dependencies:
  pip install openpyxl ortools

Example CLI usage:
  python pairwise.py --input params.xlsx --output suite.xlsx --verbose
  python pairwise.py --input params.xlsx --output suite.xlsx --no-exact
  python pairwise.py --input params.xlsx --output suite.xlsx --T_MAX 10000 --exact_time_limit 60

Input Excel format (first sheet by default):
  Row 1 = parameter names (unique, non-empty).
  Rows 2..R = possible values per column; empty cells are ignored.

Output Excel:
  Sheet "Suite" = the generated test cases (same column headers).
  Sheet "Summary" = LB, UB, FINAL_SIZE, OPTIMAL_PROVEN, T, T_MAX, etc.
"""

from __future__ import annotations

import argparse
import os
from itertools import product as cart_product
from typing import Any, Dict, List, Optional, Sequence, Tuple

import openpyxl


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    """Parse command-line arguments."""
    p = argparse.ArgumentParser(
        description="Generate a pairwise (2-way) test suite from an Excel parameter table."
    )
    p.add_argument("--input", required=True, help="Path to input Excel file")
    p.add_argument("--output", required=True, help="Path to output Excel file")
    p.add_argument("--sheet", default=None, help="Sheet name to read (default: first sheet)")
    p.add_argument("--seed", type=int, default=0, help="Deterministic seed for tie-breaking (default: 0)")
    p.add_argument("--T_MAX", type=int, default=50_000, help="Cartesian threshold for enabling exact phase (default: 50000)")
    p.add_argument("--exact_time_limit", type=float, default=None,
                   help="Optional time limit in seconds for exact phase only")
    p.add_argument("--no-exact", action="store_true", help="Force skipping the exact phase")
    p.add_argument("--verbose", action="store_true", help="Print progress logs")
    return p.parse_args(argv)


# ---------------------------------------------------------------------------
# INPUT PARSING
# ---------------------------------------------------------------------------

def read_input_excel(path: str, sheet: Optional[str] = None) -> Tuple[List[str], List[List[str]]]:
    """
    Read parameter names and their possible values from an Excel file.

    Returns:
        param_names: list of parameter name strings (column headers).
        values: list of lists; values[j] = de-duplicated list of string values
                for parameter j, preserving first-occurrence order.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], []

    # Row 1 → headers
    header_row = rows[0]
    param_names: List[str] = []
    col_indices: List[int] = []
    for col_idx, cell_val in enumerate(header_row):
        if cell_val is not None and str(cell_val).strip() != "":
            param_names.append(str(cell_val).strip())
            col_indices.append(col_idx)

    if not param_names:
        return [], []

    # Check uniqueness
    if len(param_names) != len(set(param_names)):
        raise ValueError("Parameter names must be unique.")

    # Rows 2..R → values
    data_rows = rows[1:]
    values: List[List[str]] = [[] for _ in param_names]
    for row in data_rows:
        for j, col_idx in enumerate(col_indices):
            cell = row[col_idx] if col_idx < len(row) else None
            if cell is not None and str(cell).strip() != "":
                s = str(cell).strip()
                if s not in values[j]:          # de-dup, preserve order
                    values[j].append(s)

    for j, name in enumerate(param_names):
        if len(values[j]) == 0:
            raise ValueError(f"Parameter '{name}' has no values.")

    return param_names, values


# ---------------------------------------------------------------------------
# BOUNDS
# ---------------------------------------------------------------------------

def compute_bounds(values: List[List[str]]) -> Tuple[int, int]:
    """
    Compute lower bound (LB) and Cartesian product size (T).

    LB = max(v_i * v_j) for all i < j   (0 if n < 2).
    T  = product of v_i.

    Returns (LB, T).
    """
    n = len(values)
    if n == 0:
        return 0, 0
    if n == 1:
        return len(values[0]), len(values[0])

    sizes = [len(v) for v in values]

    # T (Cartesian size)
    T = 1
    for s in sizes:
        T *= s

    # LB = max(v_i * v_j) for i < j
    LB = 0
    for i in range(n):
        for j in range(i + 1, n):
            LB = max(LB, sizes[i] * sizes[j])

    return LB, T


# ---------------------------------------------------------------------------
# REQUIREMENT UNIVERSE
# ---------------------------------------------------------------------------

def build_pair_universe(values: List[List[str]]) -> List[Tuple[int, int, str, str]]:
    """
    Build the full list of pairwise requirements.

    Each requirement is (i, j, a, b) where i < j, a ∈ values[i], b ∈ values[j].
    """
    n = len(values)
    universe: List[Tuple[int, int, str, str]] = []
    for i in range(n):
        for j in range(i + 1, n):
            for a in values[i]:
                for b in values[j]:
                    universe.append((i, j, a, b))
    return universe


# ---------------------------------------------------------------------------
# IPOG PAIRWISE SUITE
# ---------------------------------------------------------------------------

def _get_missing_pairs(
    values: List[List[str]], k: int,
) -> List[Tuple[Tuple[int, str], Tuple[int, str]]]:
    """
    Return all required pairs between parameter k and parameters 0..k-1.

    Each pair is ((i, val_i), (k, val_k)).
    """
    pairs = []
    for i in range(k):
        for vi in values[i]:
            for vk in values[k]:
                pairs.append(((i, vi), (k, vk)))
    return pairs


def _count_covered(
    row: List[str],
    val: str,
    k: int,
    missing_pairs: List[Tuple[Tuple[int, str], Tuple[int, str]]],
) -> int:
    """Count how many missing pairs would be covered by assigning row[k]=val."""
    count = 0
    for i, row_val in enumerate(row):
        if ((i, row_val), (k, val)) in missing_pairs:
            count += 1
    return count


def _remove_covered(
    row: List[str],
    k: int,
    missing_pairs: List[Tuple[Tuple[int, str], Tuple[int, str]]],
) -> None:
    """Remove from missing_pairs all pairs covered by the given row."""
    for i in range(k):
        pair = ((i, row[i]), (k, row[k]))
        if pair in missing_pairs:
            missing_pairs.remove(pair)


def greedy_pairwise_suite(
    values: List[List[str]],
    verbose: bool = False,
) -> List[Tuple[str, ...]]:
    """
    IPOG greedy generator for a valid pairwise suite.

    Algorithm (In-Parameter-Order-General):
      1. Start with the full Cartesian product of the first 2 parameters.
      2. For each subsequent parameter k (2, 3, …, n-1):
         a. HORIZONTAL GROWTH — extend every existing row with the value
            for parameter k that covers the most uncovered pairs.
         b. VERTICAL GROWTH — for any still-uncovered pair (i, k, vi, vk),
            add a new row with those forced values and fill remaining
            slots with default values.

    Returns list of test tuples (length n each).
    """
    n = len(values)
    if n == 0:
        return []
    if n == 1:
        return [(v,) for v in values[0]]

    # Step 1: seed with full Cartesian product of first 2 parameters
    suite: List[List[str]] = [
        list(combo) for combo in cart_product(values[0], values[1])
    ]

    if verbose:
        print(f"  [ipog] seed: {len(suite)} tests from params 0,1 "
              f"(sizes {len(values[0])}×{len(values[1])})")

    # Step 2: extend one parameter at a time
    for k in range(2, n):
        # Build list of all required pairs for this parameter
        missing_pairs = _get_missing_pairs(values, k)

        # --- HORIZONTAL GROWTH ---
        # Extend each existing row with the best value for parameter k
        for row in suite:
            best_val = values[k][0]
            best_gain = -1
            for val in values[k]:
                gain = _count_covered(row, val, k, missing_pairs)
                if gain > best_gain:
                    best_gain = gain
                    best_val = val
            row.append(best_val)
            _remove_covered(row, k, missing_pairs)

        if verbose:
            print(f"  [ipog] param {k}: after horizontal, "
                  f"missing={len(missing_pairs)}")

        # --- VERTICAL GROWTH ---
        # For each still-uncovered pair, add a new row
        for pair in list(missing_pairs):
            # Check if already covered by a previously added vertical row
            already_covered = any(
                row[pair[0][0]] == pair[0][1] and row[pair[1][0]] == pair[1][1]
                for row in suite
            )
            if already_covered:
                continue

            # Create new row forcing the uncovered pair
            new_row = [None] * (k + 1)
            new_row[pair[0][0]] = pair[0][1]  # force param i = val_i
            new_row[pair[1][0]] = pair[1][1]  # force param k = val_k

            # Fill remaining slots: pick value covering the most missing pairs
            for j in range(k + 1):
                if new_row[j] is not None:
                    continue
                best_fill = values[j][0]
                best_fill_gain = -1
                for val in values[j]:
                    gain = _count_covered(new_row, val, k, missing_pairs) if j != k else 0
                    # Also count pairs (j, k) covered
                    if j < k and ((j, val), (k, new_row[k])) in missing_pairs:
                        gain += 1
                    if gain > best_fill_gain:
                        best_fill_gain = gain
                        best_fill = val
                new_row[j] = best_fill

            suite.append(new_row)
            _remove_covered(new_row, k, missing_pairs)

        if verbose:
            print(f"  [ipog] param {k}: after vertical, suite={len(suite)}")

    # Convert to tuples
    return [tuple(row) for row in suite]


# ---------------------------------------------------------------------------
# CARTESIAN CANDIDATES
# ---------------------------------------------------------------------------

def build_cartesian_candidates(values: List[List[str]]) -> List[Tuple[str, ...]]:
    """Return the full Cartesian product of all parameter values."""
    return list(cart_product(*values))


# ---------------------------------------------------------------------------
# EXACT MINIMAL PHASE (CP-SAT)
# ---------------------------------------------------------------------------

def solve_exact_minimal(
    values: List[List[str]],
    candidates: List[Tuple[str, ...]],
    LB: int,
    UB: int,
    seed: int = 0,
    exact_time_limit: Optional[float] = None,
    verbose: bool = False,
) -> Tuple[Optional[List[Tuple[str, ...]]], bool, bool]:
    """
    Attempt to find the true minimal pairwise suite via CP-SAT.

    Uses the full Cartesian product as candidate tests and a single
    minimisation model: Minimize(sum(x_t)).

    Returns:
        (suite_or_None, optimal_proven, exact_used)
    """
    from ortools.sat.python import cp_model

    n = len(values)
    universe = build_pair_universe(values)

    # Pre-compute: for each requirement → list of candidate indices that cover it
    req_to_cands: Dict[Tuple[int, int, str, str], List[int]] = {r: [] for r in universe}
    for t_idx, cand in enumerate(candidates):
        for i in range(n):
            for j in range(i + 1, n):
                req = (i, j, cand[i], cand[j])
                req_to_cands[req].append(t_idx)

    # Build a single model with minimisation objective
    model = cp_model.CpModel()

    x = [model.NewBoolVar(f"x_{t}") for t in range(len(candidates))]

    # Coverage: each requirement must be covered by at least one selected candidate
    for req, cand_ids in req_to_cands.items():
        model.Add(sum(x[t] for t in cand_ids) >= 1)

    # Bound hints to help the solver
    model.Add(sum(x) >= LB)
    model.Add(sum(x) <= UB - 1)  # only interested in improvements over greedy

    # Objective: minimise total selected tests
    model.Minimize(sum(x))

    solver = cp_model.CpSolver()
    solver.parameters.num_search_workers = os.cpu_count() or 1
    solver.parameters.random_seed = seed

    if exact_time_limit is not None:
        solver.parameters.max_time_in_seconds = exact_time_limit

    if verbose:
        solver.parameters.log_search_progress = True
        print(f"  [exact] solving single model: Minimize(sum(x)), LB={LB}, UB={UB}")

    status = solver.Solve(model)

    if status in (cp_model.FEASIBLE, cp_model.OPTIMAL):
        chosen = [candidates[t] for t in range(len(candidates)) if solver.Value(x[t])]
        optimal_proven = (status == cp_model.OPTIMAL)
        if verbose:
            label = "OPTIMAL" if optimal_proven else "FEASIBLE (best found)"
            print(f"  [exact] {label} — size={len(chosen)}")
        return chosen, optimal_proven, True

    if status == cp_model.INFEASIBLE:
        if verbose:
            print("  [exact] INFEASIBLE — IPOG is already optimal")
        return None, True, True

    if verbose:
        print("  [exact] UNKNOWN/timeout — no improvement found")
    return None, False, True


# ---------------------------------------------------------------------------
# VERIFICATION
# ---------------------------------------------------------------------------

def verify_pairwise(suite: Sequence[Tuple[str, ...]], values: List[List[str]]) -> None:
    """
    Verify that `suite` achieves full pairwise coverage.
    Raises AssertionError if any pair is missing.
    """
    n = len(values)
    if n < 2:
        return  # trivially covered

    required = set(build_pair_universe(values))
    covered: set = set()

    for test in suite:
        for i in range(n):
            for j in range(i + 1, n):
                pair = (i, j, test[i], test[j])
                covered.add(pair)

    missing = required - covered
    if missing:
        raise AssertionError(
            f"Pairwise coverage FAILED — {len(missing)} pair(s) missing.\n"
            f"Examples: {list(missing)[:5]}"
        )


# ---------------------------------------------------------------------------
# OUTPUT
# ---------------------------------------------------------------------------

def write_output_excel(
    path: str,
    param_names: List[str],
    suite: List[Tuple[str, ...]],
    summary: Dict[str, Any],
) -> None:
    """
    Write the test suite and summary to an Excel workbook.

    Sheet "Suite" — test cases (same column headers).
    Sheet "Summary" — key metrics.
    """
    wb = openpyxl.Workbook()

    # ---- Suite sheet ----
    ws_suite = wb.active
    ws_suite.title = "Suite"
    for col_idx, name in enumerate(param_names, start=1):
        ws_suite.cell(row=1, column=col_idx, value=name)
    for row_idx, test in enumerate(suite, start=2):
        for col_idx, val in enumerate(test, start=1):
            ws_suite.cell(row=row_idx, column=col_idx, value=val)

    # ---- Summary sheet ----
    ws_sum = wb.create_sheet("Summary")
    summary_keys = [
        "LB", "UB", "FINAL_SIZE", "OPTIMAL_PROVEN",
        "T", "T_MAX", "EXACT_TIME_LIMIT", "EXACT_USED",
    ]
    ws_sum.cell(row=1, column=1, value="Metric")
    ws_sum.cell(row=1, column=2, value="Value")
    for row_idx, key in enumerate(summary_keys, start=2):
        ws_sum.cell(row=row_idx, column=1, value=key)
        val = summary.get(key, "")
        if isinstance(val, bool):
            val = "TRUE" if val else "FALSE"
        ws_sum.cell(row=row_idx, column=2, value="" if val is None else val)

    wb.save(path)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main(argv: Optional[List[str]] = None) -> None:
    args = parse_args(argv)

    # ---- Read input ----
    if args.verbose:
        print(f"Reading input: {args.input}")
    param_names, values = read_input_excel(args.input, args.sheet)
    n = len(param_names)

    if args.verbose:
        for j, name in enumerate(param_names):
            print(f"  {name}: {values[j]}")

    # ---- Edge cases ----
    if n == 0:
        suite: List[Tuple[str, ...]] = []
        summary: Dict[str, Any] = {
            "LB": 0, "UB": 0, "FINAL_SIZE": 0,
            "OPTIMAL_PROVEN": True, "T": 0, "T_MAX": args.T_MAX,
            "EXACT_TIME_LIMIT": args.exact_time_limit, "EXACT_USED": False,
        }
        write_output_excel(args.output, param_names, suite, summary)
        print("LB=0  UB=0  FINAL_SIZE=0  OPTIMAL_PROVEN=TRUE  (0 parameters)")
        return

    if n == 1:
        suite = [(v,) for v in values[0]]
        sz = len(suite)
        summary = {
            "LB": sz, "UB": sz, "FINAL_SIZE": sz,
            "OPTIMAL_PROVEN": True, "T": sz, "T_MAX": args.T_MAX,
            "EXACT_TIME_LIMIT": args.exact_time_limit, "EXACT_USED": False,
        }
        write_output_excel(args.output, param_names, suite, summary)
        print(f"LB={sz}  UB={sz}  FINAL_SIZE={sz}  OPTIMAL_PROVEN=TRUE  (1 parameter)")
        return

    # ---- Bounds ----
    LB, T = compute_bounds(values)
    if args.verbose:
        print(f"Bounds: LB={LB}  T={T}")

    # ---- IPOG Greedy ----
    if args.verbose:
        print("Running IPOG algorithm …")
    suite_greedy = greedy_pairwise_suite(values, verbose=args.verbose)
    UB = len(suite_greedy)
    verify_pairwise(suite_greedy, values)
    if args.verbose:
        print(f"IPOG suite verified — UB={UB}")

    # ---- Decide exact phase ----
    exact_used = False
    optimal_proven = False
    final_suite = suite_greedy
    final_size = UB

    if UB == LB:
        optimal_proven = True
        if args.verbose:
            print("IPOG already meets LB — optimal proven without exact phase.")
    elif args.no_exact:
        if args.verbose:
            print("Exact phase skipped (--no-exact).")
    elif T > args.T_MAX:
        if args.verbose:
            print(f"Exact phase skipped — Cartesian size T={T} > T_MAX={args.T_MAX}.")
    else:
        # ---- Run exact phase ----
        if args.verbose:
            print(f"Building Cartesian candidates (T={T}) …")
        candidates = build_cartesian_candidates(values)
        if args.verbose:
            print(f"Starting exact phase: Minimize(sum(x)), LB={LB}, UB={UB}")

        exact_suite, proven, _ = solve_exact_minimal(
            values, candidates, LB, UB,
            seed=args.seed,
            exact_time_limit=args.exact_time_limit,
            verbose=args.verbose,
        )
        exact_used = True

        if exact_suite is not None:
            verify_pairwise(exact_suite, values)
            final_suite = exact_suite
            final_size = len(exact_suite)
            optimal_proven = proven
            if args.verbose:
                print(f"Exact suite verified — size={final_size}, optimal_proven={optimal_proven}")
        else:
            if proven:
                optimal_proven = True
                if args.verbose:
                    print("Exact proved IPOG is optimal.")
            else:
                if args.verbose:
                    print("Exact phase did not improve on IPOG.")

    # ---- Summary ----
    skipped_reason = ""
    if T > args.T_MAX and not args.no_exact:
        skipped_reason = f"  EXACT_SKIPPED: T={T} > T_MAX={args.T_MAX}"

    summary = {
        "LB": LB, "UB": UB, "FINAL_SIZE": final_size,
        "OPTIMAL_PROVEN": optimal_proven,
        "T": T, "T_MAX": args.T_MAX,
        "EXACT_TIME_LIMIT": args.exact_time_limit,
        "EXACT_USED": exact_used,
    }

    # ---- Write output ----
    write_output_excel(args.output, param_names, final_suite, summary)
    if args.verbose:
        print(f"Output written to {args.output}")

    # ---- Stdout report ----
    print(
        f"LB={LB}  UB={UB}  FINAL_SIZE={final_size}  "
        f"OPTIMAL_PROVEN={'TRUE' if optimal_proven else 'FALSE'}"
        f"{skipped_reason}"
    )


if __name__ == "__main__":
    main()
